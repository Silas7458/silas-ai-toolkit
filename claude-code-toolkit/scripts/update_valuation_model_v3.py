"""
Hospice Valuation Model v3.0 Update Script
===========================================
Updates hospice_valuation_model_2_5_8.xlsx -> hospice_valuation_model_3_0.xlsx

Changes:
1. Recalibrate auto-estimate multiples to 7-tier ADC framework (was 3-tier)
2. NOI multiple becomes dynamic (65% of EBITDA-A EV, translated to NOI basis)
3. Add Market & Sensitivity Adjustments section (rows 89-102)
4. Add Sensitivity Scoring sheet (19 factors, weighted, linked to Dashboard)
5. Add ADC Tier Reference Table on Sensitivity Scoring sheet
6. Update version label to v3.0
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# Paths
SRC = '{{HOME_DIR}}/Downloads/hospice_valuation_model_2_5_8.xlsx'
DST = '{{HOME_DIR}}/Downloads/hospice_valuation_model_3_0.xlsx'

# Load workbook (preserve formulas)
wb = openpyxl.load_workbook(SRC)
ws = wb['Valuation Dashboard']

# =================================================================
# STYLE DEFINITIONS (match existing model)
# =================================================================
DARK_BLUE = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
MED_BLUE = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
LIGHT_BLUE = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
PINK = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
GREEN = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
YELLOW = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
INPUT_BG = PatternFill(start_color='FFFFF0', end_color='FFFFF0', fill_type='solid')

HEADER = Font(name='Arial', size=12, bold=True, color='FFFFFF')
SUBHEADER = Font(name='Arial', size=10, bold=True, color='FFFFFF')
LABEL = Font(name='Arial', size=11)
LABEL_BOLD = Font(name='Arial', size=11, bold=True)
VALUE = Font(name='Arial', size=11, bold=True)
NOTE = Font(name='Arial', size=9, color='666666')
BIG_VALUE = Font(name='Arial', size=14, bold=True, color='2F5496')

THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
MEDIUM_BORDER = Border(
    left=Side(style='medium', color='2F5496'),
    right=Side(style='medium', color='2F5496'),
    top=Side(style='medium', color='2F5496'),
    bottom=Side(style='medium', color='2F5496')
)


def section_header(sheet, row, text):
    """Create a dark blue section header merged across B:H"""
    sheet.merge_cells(f'B{row}:H{row}')
    cell = sheet[f'B{row}']
    cell.value = text
    cell.font = HEADER
    cell.fill = DARK_BLUE
    cell.alignment = Alignment(horizontal='left', vertical='center')
    for c in 'BCDEFGH':
        sheet[f'{c}{row}'].border = THIN
        sheet[f'{c}{row}'].fill = DARK_BLUE


def sub_headers(sheet, row, pairs):
    """Create sub-header row with column:label pairs"""
    for col, label in pairs:
        cell = sheet[f'{col}{row}']
        cell.value = label
        cell.font = SUBHEADER
        cell.fill = MED_BLUE
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN


# =================================================================
# 1. VERSION LABEL
# =================================================================
ws['B2'] = 'Hospice Enterprise Valuation Model v3.0'
print('[1/6] Version label updated')

# =================================================================
# 2. RECALIBRATE AUTO-ESTIMATE FORMULAS
# =================================================================

# EBITDA-A (D60): 7-tier ADC + margin bonus
# Tiers: <10=0, 10-20=4.0, 20-50=4.7, 50-80=6.0, 80-120=7.5, 120-200=10.0, >200=14.5
# Margin bonus: >25%=+1.0, >20%=+0.5, else 0
ws['D60'] = (
    '=IF(D30<10,0,'
    'IF(D30<20,4,'
    'IF(D30<50,4.7,'
    'IF(D30<80,6,'
    'IF(D30<120,7.5,'
    'IF(D30<200,10,14.5))))))'
    '+IF(H28>0.25,1,IF(H28>0.2,0.5,0))'
)

# NOI / EBITDA-B (D59): Dynamic — targets 65% of EBITDA-A EV
# Translates to NOI basis: EBITDA_mult * (EBITDA/NOI) * 0.65
ws['D59'] = '=IF(D48<=0,0,ROUND(H60*(D44/D48)*0.65,1))'

# Revenue (D61): 7-tier ADC + margin bonus
# Tiers: <10=0.5, 10-20=0.6, 20-50=0.8, 50-80=1.0, 80-120=1.3, 120-200=1.8, >200=2.2
ws['D61'] = (
    '=IF(D30<10,0.5,'
    'IF(D30<20,0.6,'
    'IF(D30<50,0.8,'
    'IF(D30<80,1,'
    'IF(D30<120,1.3,'
    'IF(D30<200,1.8,2.2))))))'
    '+IF(H28>0.25,0.15,IF(H28>0.2,0.1,0))'
)

# SDE (D62): 7-tier ADC + margin bonus
# Tiers: <10=0, 10-20=2.5, 20-50=3.5, 50-80=4.0, 80-120=4.5, 120-200=5.0, >200=6.0
ws['D62'] = (
    '=IF(D30<10,0,'
    'IF(D30<20,2.5,'
    'IF(D30<50,3.5,'
    'IF(D30<80,4,'
    'IF(D30<120,4.5,'
    'IF(D30<200,5,6))))))'
    '+IF(H28>0.25,0.5,IF(H28>0.2,0.25,0))'
)

# Notes column
for r, text in [
    (59, 'Auto: 65% of EBITDA-A EV, as NOI multiple'),
    (60, '7-tier: 0/4.0/4.7/6.0/7.5/10.0/14.5 + margin'),
    (61, '7-tier: 0.5/0.6/0.8/1.0/1.3/1.8/2.2 + margin'),
    (62, '7-tier: 0/2.5/3.5/4.0/4.5/5.0/6.0 + margin'),
]:
    ws[f'J{r}'] = text
    ws[f'J{r}'].font = NOTE

# Update section header
ws['B57'] = 'Valuation Multiples \u2014 7-Tier ADC Auto-Estimates (Texas Hospice Market)'

print('[2/6] Auto-estimate formulas recalibrated to 7-tier ADC framework')

# =================================================================
# 3. MARKET & SENSITIVITY ADJUSTMENTS (Rows 89-93)
# =================================================================
section_header(ws, 89, 'Market & Sensitivity Adjustments')
sub_headers(ws, 90, [
    ('B', 'Factor'), ('C', 'Input'), ('D', 'Auto-Estimate'),
    ('F', 'Override'), ('H', 'Applied')
])

# Row 91: Non-CON State toggle
ws['B91'] = 'Non-CON State (e.g., Texas)'
ws['B91'].font = LABEL
ws['C91'] = 'yes'
ws['C91'].font = VALUE
ws['C91'].fill = INPUT_BG
ws['C91'].alignment = Alignment(horizontal='center')

# Auto TX discount by ADC tier (research-derived)
# <10: -16%, 10-20: -12.5%, 20-50: -7%, 50-80: -6%, 80-120: -4%, 120-200: -1%, >200: +1%
ws['D91'] = (
    '=IF(C91="yes",'
    'IF(D30<10,-0.16,'
    'IF(D30<20,-0.125,'
    'IF(D30<50,-0.07,'
    'IF(D30<80,-0.06,'
    'IF(D30<120,-0.04,'
    'IF(D30<200,-0.01,0.01)))))),'
    '0)'
)
ws['D91'].font = VALUE
ws['D91'].number_format = '0.0%'
ws['H91'] = '=IF(F91="",D91,F91)'
ws['H91'].font = VALUE
ws['H91'].number_format = '0.0%'
ws['H91'].fill = LIGHT_BLUE
ws['J91'] = 'TX non-CON discount by ADC tier (Scope Research, Stoneridge)'
ws['J91'].font = NOTE

# Row 92: Sensitivity Composite Score (linked to Sensitivity Scoring sheet)
ws['B92'] = 'Sensitivity Composite Score'
ws['B92'].font = LABEL
# C92 will be set after we know the total row on Sensitivity Scoring sheet
# Placeholder — overwritten below after sheet creation
ws['C92'] = 0
ws['C92'].font = VALUE
ws['C92'].fill = INPUT_BG
ws['C92'].alignment = Alignment(horizontal='center')

# Convert composite score to EV impact %: (score * 0.046) / base EBITDA multiple
ws['D92'] = '=IF(H60=0,0,(C92*0.046)/H60)'
ws['D92'].font = VALUE
ws['D92'].number_format = '0.0%'
ws['J92'] = 'Linked to Sensitivity Scoring sheet (or enter score manually in C92)'
ws['J92'].font = NOTE

# Row 93: Combined Market Adjustment
ws['B93'] = 'Combined Market Adjustment'
ws['B93'].font = LABEL_BOLD
ws['H93'] = '=H91+D92'
ws['H93'].font = VALUE
ws['H93'].number_format = '0.0%'
ws['H93'].fill = YELLOW
ws['H93'].border = MEDIUM_BORDER
ws['J93'] = 'State adj + Sensitivity adj (applied to entire EV range)'
ws['J93'].font = NOTE

# Data validation: C91 must be "yes" or "no"
dv_yn = DataValidation(type='list', formula1='"yes,no"', allow_blank=False)
dv_yn.error = 'Enter yes or no'
dv_yn.errorTitle = 'Invalid Input'
ws.add_data_validation(dv_yn)
dv_yn.add(ws['C91'])

print('[3/6] Market & Sensitivity Adjustments section added (rows 89-93)')

# =================================================================
# 4. FINAL MARKET-ADJUSTED VALUATION (Rows 95-102)
# =================================================================
section_header(ws, 95, 'Final Market-Adjusted Enterprise Value')
sub_headers(ws, 96, [('B', 'Measure'), ('H', 'Enterprise Value')])

# Market-adjusted range (applies combined factor to CAP-adjusted range)
for row, label, fill, src_row in [
    (97, 'Market-Adjusted Low', PINK, 82),
    (98, 'Market-Adjusted Mid', LIGHT_BLUE, 83),
    (99, 'Market-Adjusted High', GREEN, 84),
]:
    ws[f'B{row}'] = label
    ws[f'B{row}'].font = LABEL
    ws[f'H{row}'] = f'=H{src_row}*(1+H93)'
    ws[f'H{row}'].font = VALUE
    ws[f'H{row}'].number_format = '$#,##0'
    ws[f'H{row}'].fill = fill

# Final recommended value
section_header(ws, 101, 'Recommended Enterprise Value')

ws['B102'] = 'Final Market-Adjusted Value'
ws['B102'].font = LABEL_BOLD
ws['H102'] = '=AVERAGE(H97:H99)'
ws['H102'].font = BIG_VALUE
ws['H102'].number_format = '$#,##0'
ws['H102'].fill = LIGHT_BLUE
ws['H102'].border = MEDIUM_BORDER

# Also add a per-ADC reference for quick sanity check
ws['B103'] = 'Implied Value Per ADC'
ws['B103'].font = LABEL
ws['H103'] = '=IF(D30=0,0,H102/D30)'
ws['H103'].font = VALUE
ws['H103'].number_format = '$#,##0'
ws['J103'] = 'Final EV / ADC (cross-check: $45K-$70K for 20-50 ADC tier)'
ws['J103'].font = NOTE

print('[4/6] Final Market-Adjusted Valuation section added (rows 95-103)')

# =================================================================
# 5. SENSITIVITY SCORING SHEET
# =================================================================
ws2 = wb.create_sheet('Sensitivity Scoring')

# Title
ws2['B2'] = 'Hospice Valuation \u2014 19-Factor Sensitivity Scoring'
ws2['B2'].font = Font(name='Arial', size=14, bold=True, color='2F5496')
ws2['B3'] = 'Score each factor from -2 (worst) to +2 (best). Weighted total feeds into Valuation Dashboard automatically.'
ws2['B3'].font = Font(name='Arial', size=10, color='666666')
ws2['B4'] = 'Multiple Adjustment = Composite Score \u00d7 0.046 (additive to EBITDA multiple, converted to % for EV application)'
ws2['B4'].font = Font(name='Arial', size=9, color='999999')

# Column headers (Row 5)
sub_headers(ws2, 5, [
    ('B', 'Factor'), ('C', 'Tier'), ('D', 'Weight'),
    ('E', 'Score (-2 to +2)'), ('F', 'Weighted Score'), ('G', 'Scoring Guide')
])

# 19 Sensitivity Factors organized by tier
FACTORS = [
    # Tier 1 — Large Impact (Weight 3)
    ('Revenue / EBITDA Trend', 1, '+2=strong growth, 0=flat, -2=declining'),
    ('Payer Mix Quality', 1, '+2=high Medicare, 0=average, -2=heavy Medicaid'),
    ('Survey / Compliance History', 1, '+2=clean surveys, 0=minor findings, -2=sanctions/CMPs'),
    ('EBITDA Margin vs Peers', 1, '+2= >20%, 0=12-15%, -2= <8%'),
    ('Operator Quality / Transferability', 1, '+2=strong team stays, 0=adequate, -2=key-person risk'),
    ('CON / License Protection', 1, '+2=CON state, 0=non-CON, -2=enhanced oversight state'),
    ('Geographic Market Strength', 1, '+2=high-growth metro, 0=average, -2=rural/saturated'),
    # Tier 2 — Moderate Impact (Weight 2)
    ('Average Length of Stay (ALOS)', 2, '+2=optimal 60-90d, 0=30-60d, -2= <20d or >120d'),
    ('Referral Concentration Risk', 2, '+2=diverse sources, 0=moderate, -2=single source >40%'),
    ('Staff Retention / Turnover', 2, '+2= <15% turnover, 0=15-25%, -2= >40%'),
    ('Live Discharge Rate', 2, '+2= <10%, 0=10-20%, -2= >25%'),
    ('CAHPS / Quality Scores', 2, '+2=4+ stars, 0=3 stars, -2= <2 stars'),
    ('Payer Contract Breadth (MA)', 2, '+2=multiple MA contracts, 0=some, -2=FFS only'),
    # Tier 3 — Incremental Impact (Weight 1)
    ('EMR / Technology Stack', 3, '+2=modern cloud EMR, 0=adequate, -2=paper/legacy'),
    ('Facility / Lease Terms', 3, '+2=favorable long-term, 0=OK, -2=expiring/unfavorable'),
    ('Brand / Community Reputation', 3, '+2=strong local brand, 0=neutral, -2=negative'),
    ('De Novo vs Mature Operations', 3, '+2= 5+ years, 0=2-5 years, -2= <2 years'),
    ('Pending Regulatory Changes', 3, '+2=favorable outlook, 0=neutral, -2=adverse pending'),
    ('Documentation / HOPE Readiness', 3, '+2=HOPE-ready, 0=partial, -2=significant gaps'),
]

TIER_COLORS = {
    1: PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid'),
    2: PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid'),
    3: PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid'),
}
TIER_LABELS = {
    1: 'Tier 1 \u2014 Large Impact (Weight 3\u00d7)',
    2: 'Tier 2 \u2014 Moderate Impact (Weight 2\u00d7)',
    3: 'Tier 3 \u2014 Incremental Impact (Weight 1\u00d7)',
}
WEIGHT_MAP = {1: 3, 2: 2, 3: 1}

current_tier = 0
data_rows = []
row = 6

for name, tier, guide in FACTORS:
    # Insert tier section header when tier changes
    if tier != current_tier:
        ws2.merge_cells(f'B{row}:G{row}')
        ws2[f'B{row}'] = TIER_LABELS[tier]
        ws2[f'B{row}'].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        for c in 'BCDEFG':
            ws2[f'{c}{row}'].fill = MED_BLUE
            ws2[f'{c}{row}'].border = THIN
        current_tier = tier
        row += 1

    weight = WEIGHT_MAP[tier]

    ws2[f'B{row}'] = name
    ws2[f'B{row}'].font = LABEL

    ws2[f'C{row}'] = tier
    ws2[f'C{row}'].font = VALUE
    ws2[f'C{row}'].alignment = Alignment(horizontal='center')
    ws2[f'C{row}'].fill = TIER_COLORS[tier]

    ws2[f'D{row}'] = weight
    ws2[f'D{row}'].font = VALUE
    ws2[f'D{row}'].alignment = Alignment(horizontal='center')

    ws2[f'E{row}'] = 0
    ws2[f'E{row}'].font = VALUE
    ws2[f'E{row}'].alignment = Alignment(horizontal='center')
    ws2[f'E{row}'].fill = INPUT_BG

    ws2[f'F{row}'] = f'=D{row}*E{row}'
    ws2[f'F{row}'].font = VALUE
    ws2[f'F{row}'].alignment = Alignment(horizontal='center')

    ws2[f'G{row}'] = guide
    ws2[f'G{row}'].font = NOTE

    data_rows.append(row)
    row += 1

# Composite Score Total
total_row = row + 1
first_data = data_rows[0]
last_data = data_rows[-1]

ws2.merge_cells(f'B{total_row}:D{total_row}')
ws2[f'B{total_row}'] = 'COMPOSITE SCORE'
ws2[f'B{total_row}'].font = BIG_VALUE

ws2[f'F{total_row}'] = f'=SUM(F{first_data}:F{last_data})'
ws2[f'F{total_row}'].font = BIG_VALUE
ws2[f'F{total_row}'].alignment = Alignment(horizontal='center')
ws2[f'F{total_row}'].fill = LIGHT_BLUE
ws2[f'F{total_row}'].border = MEDIUM_BORDER

ws2[f'G{total_row}'] = f'Range: -{7*3*2 + 6*2*2 + 6*1*2} to +{7*3*2 + 6*2*2 + 6*1*2}'
ws2[f'G{total_row}'].font = NOTE

# Info rows
info_row = total_row + 2
ws2[f'B{info_row}'] = 'Multiple Adjustment = Composite Score \u00d7 0.046'
ws2[f'B{info_row}'].font = NOTE
ws2[f'B{info_row+1}'] = f'This total (F{total_row}) is linked to cell C92 on the Valuation Dashboard'
ws2[f'B{info_row+1}'].font = Font(name='Arial', size=10, bold=True, color='2F5496')

# Data validation: scores must be -2 to +2
dv_score = DataValidation(type='whole', operator='between', formula1=-2, formula2=2)
dv_score.error = 'Score must be between -2 and +2'
dv_score.errorTitle = 'Invalid Score'
ws2.add_data_validation(dv_score)
for r in data_rows:
    dv_score.add(ws2[f'E{r}'])

# Column widths
ws2.column_dimensions['B'].width = 36
ws2.column_dimensions['C'].width = 8
ws2.column_dimensions['D'].width = 10
ws2.column_dimensions['E'].width = 18
ws2.column_dimensions['F'].width = 16
ws2.column_dimensions['G'].width = 60

print(f'[5/6] Sensitivity Scoring sheet created ({len(FACTORS)} factors, total at F{total_row})')

# =================================================================
# 5b. ADC TIER REFERENCE TABLE (on Sensitivity Scoring sheet)
# =================================================================
ref_start = info_row + 3

ws2.merge_cells(f'B{ref_start}:G{ref_start}')
ws2[f'B{ref_start}'] = '7-Tier ADC Multiple Reference Table'
ws2[f'B{ref_start}'].font = HEADER
for c in 'BCDEFG':
    ws2[f'{c}{ref_start}'].fill = DARK_BLUE
    ws2[f'{c}{ref_start}'].border = THIN

# Table headers
ref_hdr = ref_start + 1
for col, label in [('B', 'ADC Tier'), ('C', 'ADC Range'), ('D', 'EBITDA-A'), ('E', 'Revenue'), ('F', 'SDE'), ('G', 'TX Discount')]:
    ws2[f'{col}{ref_hdr}'] = label
    ws2[f'{col}{ref_hdr}'].font = SUBHEADER
    ws2[f'{col}{ref_hdr}'].fill = MED_BLUE
    ws2[f'{col}{ref_hdr}'].alignment = Alignment(horizontal='center')
    ws2[f'{col}{ref_hdr}'].border = THIN

# Table data
TIERS = [
    ('Startup', '<10', 'N/A', '0.5x', 'N/A', '-16%'),
    ('Emerging', '10-20', '4.0x', '0.6x', '2.5x', '-12.5%'),
    ('Small Profitable', '20-50', '4.7x', '0.8x', '3.5x', '-7.0%'),
    ('Medium', '50-80', '6.0x', '1.0x', '4.0x', '-6.0%'),
    ('Larger Medium', '80-120', '7.5x', '1.3x', '4.5x', '-4.0%'),
    ('Large', '120-200', '10.0x', '1.8x', '5.0x', '-1.0%'),
    ('Enterprise', '>200', '14.5x', '2.2x', '6.0x', '+1.0%'),
]

for i, (tier_name, adc_range, ebitda, rev, sde, tx_disc) in enumerate(TIERS):
    r = ref_hdr + 1 + i
    ws2[f'B{r}'] = tier_name
    ws2[f'B{r}'].font = LABEL
    ws2[f'C{r}'] = adc_range
    ws2[f'C{r}'].font = VALUE
    ws2[f'C{r}'].alignment = Alignment(horizontal='center')
    ws2[f'D{r}'] = ebitda
    ws2[f'D{r}'].font = VALUE
    ws2[f'D{r}'].alignment = Alignment(horizontal='center')
    ws2[f'E{r}'] = rev
    ws2[f'E{r}'].font = VALUE
    ws2[f'E{r}'].alignment = Alignment(horizontal='center')
    ws2[f'F{r}'] = sde
    ws2[f'F{r}'].font = VALUE
    ws2[f'F{r}'].alignment = Alignment(horizontal='center')
    ws2[f'G{r}'] = tx_disc
    ws2[f'G{r}'].font = VALUE
    ws2[f'G{r}'].alignment = Alignment(horizontal='center')
    # Highlight the Small Profitable row (our ADC 40 target)
    if tier_name == 'Small Profitable':
        for c in 'BCDEFG':
            ws2[f'{c}{r}'].fill = YELLOW

# Source note
src_row = ref_hdr + 1 + len(TIERS) + 1
ws2[f'B{src_row}'] = 'Sources: Scope Research (73 deals), Mertz Taggart, Vallexa, FOCUS Bankers, Stoneridge, HealthFMV, Hospice News'
ws2[f'B{src_row}'].font = Font(name='Arial', size=8, color='999999')

# =================================================================
# 6. LINK SENSITIVITY SCORE TO DASHBOARD
# =================================================================
ws['C92'] = f"='Sensitivity Scoring'!F{total_row}"
ws['C92'].font = VALUE
ws['C92'].fill = INPUT_BG
ws['C92'].alignment = Alignment(horizontal='center')

# =================================================================
# SAVE
# =================================================================
wb.save(DST)

print(f'[6/6] Model saved to: {DST}')
print()
print('=== v3.0 CHANGE SUMMARY ===')
print(f'  EBITDA-A multiple: 6.5x -> 4.7x (ADC 40 tier)')
print(f'  NOI multiple: 8.5x -> dynamic (~9.2x at current values)')
print(f'  Revenue multiple: 0.8x (unchanged)')
print(f'  SDE multiple: 3.5x (unchanged)')
print(f'  New: Market & Sensitivity Adjustments section (rows 89-103)')
print(f'  New: Sensitivity Scoring sheet (19 factors, F{total_row} = composite)')
print(f'  New: Final Market-Adjusted EV with TX discount (rows 95-103)')
print(f'  New: ADC Tier Reference Table on Sensitivity Scoring sheet')
